"""Attendance routes — real GPS check + 10-min edit window."""
import math
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session as DBSession

from core.security import get_current_user, require_roles
from db.database import get_db
from models.models import (AttendanceMethod, AttendanceRecord, AttendanceStatus,
                            Session, SessionStatus, SystemSettings, User, UserRole)
from schemas.schemas import AttendanceListOut, AttendanceMarkManual, AttendanceMarkQR, AttendanceOut

router = APIRouter()


def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    p1, p2 = math.radians(float(lat1)), math.radians(float(lat2))
    dp = math.radians(float(lat2)-float(lat1))
    dl = math.radians(float(lon2)-float(lon1))
    a  = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))


def get_settings(db):
    s = db.query(SystemSettings).filter(SystemSettings.id == 1).first()
    return s or SystemSettings()


def check_edit_window(session: Session, db: DBSession):
    """Raises 403 if manual edit window has passed."""
    if session.status == SessionStatus.active:
        return  # still live — always editable
    if session.status == SessionStatus.closed and session.ended_at:
        settings = get_settings(db)
        window = settings.manual_edit_window if hasattr(settings, 'manual_edit_window') else 10
        deadline = session.ended_at + timedelta(minutes=window)
        if datetime.utcnow() > deadline:
            raise HTTPException(
                status_code=403,
                detail=f"Edit window closed. Attendance can only be edited up to {window} minutes after session ends."
            )


@router.post("/qr-gps-face", response_model=AttendanceOut, status_code=201)
def mark_full_flow(
    payload: AttendanceMarkQR,
    current_user: User = Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    session = db.query(Session).filter(
        Session.qr_token == payload.qr_token,
        Session.status   == SessionStatus.active
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Invalid or expired QR code.")

    # Sub-section check — if session has a sub_section set (lab batch),
    # only students whose course matches that sub_section can mark attendance
    if session.sub_section:
        student_subsec = (current_user.course or "").strip().upper()
        session_subsec = session.sub_section.strip().upper()
        if student_subsec != session_subsec:
            raise HTTPException(
                status_code=403,
                detail=f"This lab session is for batch {session.sub_section} only. Your batch is {current_user.course or 'not set'}."
            )

    # GPS check — server-side
    if session.gps_lat and session.gps_lng:
        if not payload.student_lat or not payload.student_lng:
            raise HTTPException(status_code=400, detail="GPS coordinates required.")
        try:
            dist    = haversine(session.gps_lat, session.gps_lng,
                                payload.student_lat, payload.student_lng)
            allowed = get_settings(db).gps_range or 50
            if dist > allowed:
                raise HTTPException(
                    status_code=403,
                    detail=f"You are {int(dist)}m away. Must be within {allowed}m of classroom."
                )
        except HTTPException:
            raise
        except Exception:
            pass

    existing = db.query(AttendanceRecord).filter(
        AttendanceRecord.session_id == session.id,
        AttendanceRecord.student_id == current_user.id
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Attendance already marked.")

    att_status = AttendanceStatus.present
    if session.started_at:
        if datetime.utcnow() > session.started_at + timedelta(minutes=session.grace_minutes):
            att_status = AttendanceStatus.late

    record = AttendanceRecord(
        session_id  = session.id,
        student_id  = current_user.id,
        method      = AttendanceMethod.qr_gps_face,
        status      = att_status,
        student_lat = payload.student_lat,
        student_lng = payload.student_lng,
    )
    db.add(record); db.commit(); db.refresh(record)
    return record


@router.post("/manual", response_model=AttendanceOut, status_code=201)
def mark_manual(
    payload: AttendanceMarkManual,
    current_user: User = Depends(require_roles(UserRole.faculty, UserRole.admin)),
    db: DBSession = Depends(get_db),
):
    session = db.query(Session).filter(Session.id == payload.session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found.")

    # Check faculty owns this session
    if current_user.role == UserRole.faculty and session.faculty_id != current_user.id:
        raise HTTPException(status_code=403, detail="This session belongs to another faculty.")

    # Check 10-minute edit window
    check_edit_window(session, db)

    existing = db.query(AttendanceRecord).filter(
        AttendanceRecord.session_id == payload.session_id,
        AttendanceRecord.student_id == payload.student_id
    ).first()
    if existing:
        existing.status = payload.status
        existing.method = AttendanceMethod.manual
        existing.notes  = payload.notes
        db.commit(); db.refresh(existing)
        return existing

    record = AttendanceRecord(
        session_id=payload.session_id, student_id=payload.student_id,
        method=AttendanceMethod.manual, status=payload.status, notes=payload.notes,
    )
    db.add(record); db.commit(); db.refresh(record)
    return record


@router.get("/session/{session_id}", response_model=AttendanceListOut)
def session_attendance(
    session_id: int,
    _: User = Depends(require_roles(UserRole.faculty, UserRole.admin)),
    db: DBSession = Depends(get_db),
):
    records = db.query(AttendanceRecord).filter(AttendanceRecord.session_id == session_id).all()
    return {"total": len(records), "records": records}


@router.get("/student/{student_id}", response_model=AttendanceListOut)
def student_history(
    student_id: int,
    course_id:  Optional[int] = None,
    skip: int = 0, limit: int = 200,
    current_user: User = Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    if current_user.role == UserRole.student and current_user.id != student_id:
        raise HTTPException(status_code=403, detail="Access denied.")
    q = db.query(AttendanceRecord).filter(AttendanceRecord.student_id == student_id)
    if course_id:
        q = q.join(Session, AttendanceRecord.session_id == Session.id).filter(Session.course_id == course_id)
    total   = q.count()
    records = q.order_by(AttendanceRecord.marked_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "records": records}


# ── Feature 1: Excel / CSV Export ─────────────────────────────────────────────
import io, csv
from fastapi import Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import joinedload
from models.models import Course

@router.get("/export/session/{session_id}")
def export_session_attendance(
    session_id: int,
    format: str = Query("excel", regex="^(excel|csv)$"),
    current_user: User = Depends(require_roles(UserRole.faculty, UserRole.admin)),
    db: DBSession = Depends(get_db),
):
    """Export attendance for a session as Excel or CSV."""
    session = db.query(Session).filter(Session.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found.")

    # Get all students in this section/branch
    q = db.query(User).filter(User.role == UserRole.student, User.status == "active")
    if session.branch:  q = q.filter(User.branch  == session.branch)
    if session.section: q = q.filter(User.section == session.section)
    students = q.order_by(User.full_name).all()

    # Get attendance records
    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.session_id == session_id
    ).all()
    marked = {r.student_id: r for r in records}

    # Get course name
    course = db.query(Course).filter(Course.id == session.course_id).first()

    rows = []
    for i, stu in enumerate(students, 1):
        rec = marked.get(stu.id)
        rows.append({
            "S.No":           i,
            "Enrollment No.": stu.inst_id,
            "Student Name":   stu.full_name,
            "Branch":         stu.branch or "",
            "Section":        stu.section or "",
            "Status":         rec.status.value.upper() if rec else "ABSENT",
            "Method":         rec.method.value if rec else "-",
            "Marked At":      rec.marked_at.strftime("%d-%b-%Y %H:%M") if rec else "-",
        })

    filename = f"attendance_{course.code if course else 'session'}_{session_id}"

    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=rows[0].keys() if rows else [])
        writer.writeheader()
        writer.writerows(rows)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
    else:
        # Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Title row
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"Attendance Report — {course.name if course else 'Session'} | {session.title or ''}"
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center")

        # Info row
        ws.merge_cells("A2:H2")
        ws["A2"].value = f"Date: {session.scheduled_at.strftime('%d %B %Y')} | Branch: {session.branch or '-'} | Section: {session.section or '-'} | Total Students: {len(students)}"
        ws["A2"].font = Font(size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Header row
        headers = list(rows[0].keys()) if rows else []
        header_fill   = PatternFill("solid", fgColor="1a3c6e")
        header_font   = Font(bold=True, color="FFFFFF", size=11)
        thin_border   = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Data rows
        present_fill = PatternFill("solid", fgColor="d1fae5")
        absent_fill  = PatternFill("solid", fgColor="fee2e2")
        late_fill    = PatternFill("solid", fgColor="fef3c7")

        for row_idx, row in enumerate(rows, 4):
            for col_idx, (key, val) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_idx in [1,4,5,6,7,8] else "left")
                if key == "Status":
                    if val == "PRESENT": cell.fill = present_fill
                    elif val == "ABSENT": cell.fill = absent_fill
                    elif val == "LATE":  cell.fill = late_fill

        # Column widths
        col_widths = [6, 18, 28, 20, 10, 12, 14, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Summary below data
        summary_row = len(rows) + 5
        present_count = sum(1 for r in rows if r["Status"] == "PRESENT")
        absent_count  = sum(1 for r in rows if r["Status"] == "ABSENT")
        late_count    = sum(1 for r in rows if r["Status"] == "LATE")
        ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=f"Present: {present_count}")
        ws.cell(row=summary_row, column=3, value=f"Absent: {absent_count}")
        ws.cell(row=summary_row, column=4, value=f"Late: {late_count}")
        ws.cell(row=summary_row, column=5, value=f"Total: {len(rows)}")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )


@router.get("/export/student/{student_id}")
def export_student_attendance(
    student_id: int,
    format: str = Query("excel", regex="^(excel|csv)$"),
    current_user: User = Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    """Export a student's complete attendance record."""
    if current_user.role == UserRole.student and current_user.id != student_id:
        raise HTTPException(status_code=403)

    student = db.query(User).filter(User.id == student_id).first()
    if not student: raise HTTPException(status_code=404)

    records = db.query(AttendanceRecord)\
        .filter(AttendanceRecord.student_id == student_id)\
        .order_by(AttendanceRecord.marked_at.desc()).all()

    rows = []
    for i, rec in enumerate(records, 1):
        sess   = db.query(Session).filter(Session.id == rec.session_id).first()
        course = db.query(Course).filter(Course.id == sess.course_id).first() if sess else None
        rows.append({
            "S.No":       i,
            "Date":       rec.marked_at.strftime("%d-%b-%Y") if rec.marked_at else "-",
            "Course":     course.name if course else "-",
            "Code":       course.code if course else "-",
            "Session":    sess.title if sess else "-",
            "Status":     rec.status.value.upper(),
            "Method":     rec.method.value,
            "Time":       rec.marked_at.strftime("%H:%M") if rec.marked_at else "-",
        })

    filename = f"attendance_{student.inst_id}_{student.full_name.replace(' ','_')}"

    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=rows[0].keys() if rows else [])
        writer.writeheader()
        writer.writerows(rows)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
    else:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Attendance"

        ws.merge_cells("A1:H1")
        ws["A1"].value = f"Attendance Report — {student.full_name} ({student.inst_id})"
        ws["A1"].font  = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:H2")
        present = sum(1 for r in rows if r["Status"] == "PRESENT")
        total   = len(rows)
        pct     = round((present / total) * 100) if total else 0
        ws["A2"].value = f"Branch: {student.branch or '-'} | Section: {student.section or '-'} | Overall: {present}/{total} ({pct}%)"
        ws["A2"].font  = Font(size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = list(rows[0].keys()) if rows else []
        header_fill = PatternFill("solid", fgColor="1a3c6e")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        present_fill = PatternFill("solid", fgColor="d1fae5")
        absent_fill  = PatternFill("solid", fgColor="fee2e2")
        late_fill    = PatternFill("solid", fgColor="fef3c7")
        for row_idx, row in enumerate(rows, 4):
            for col_idx, (key, val) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if key == "Status":
                    if val == "PRESENT": cell.fill = present_fill
                    elif val == "ABSENT": cell.fill = absent_fill
                    elif val == "LATE":  cell.fill = late_fill

        col_widths = [6, 14, 30, 10, 25, 10, 14, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )


# ── Feature 2: AI Insights ────────────────────────────────────────────────────
@router.get("/insights/student/{student_id}")
def student_insights(
    student_id: int,
    current_user: User = Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    """AI-style attendance insights for a student."""
    if current_user.role == UserRole.student and current_user.id != student_id:
        raise HTTPException(status_code=403)

    student = db.query(User).filter(User.id == student_id).first()
    if not student: raise HTTPException(status_code=404)

    records = db.query(AttendanceRecord)\
        .filter(AttendanceRecord.student_id == student_id).all()

    sessions_attended = [r for r in records if r.status.value == "present"]
    total_sessions    = db.query(Session).filter(
        Session.branch  == student.branch,
        Session.section == student.section,
        Session.status  == "closed"
    ).count()

    present_count = len(sessions_attended)
    pct = round((present_count / total_sessions) * 100) if total_sessions else 0

    # Classes needed to reach 75%
    needed = 0
    if pct < 75 and total_sessions > 0:
        # solve: (present + x) / (total + x) >= 0.75
        x = 0
        while True:
            if total_sessions + x == 0: break
            if (present_count + x) / (total_sessions + x) >= 0.75:
                needed = x
                break
            x += 1
            if x > 200: needed = -1; break

    # Day-wise analysis
    from collections import defaultdict
    day_stats = defaultdict(lambda: {"present": 0, "total": 0})
    for rec in records:
        sess = db.query(Session).filter(Session.id == rec.session_id).first()
        if sess and sess.scheduled_at:
            day = sess.scheduled_at.strftime("%A")
            day_stats[day]["total"] += 1
            if rec.status.value == "present":
                day_stats[day]["present"] += 1

    worst_day = None
    worst_pct = 100
    for day, stat in day_stats.items():
        if stat["total"] > 0:
            dp = round((stat["present"] / stat["total"]) * 100)
            if dp < worst_pct:
                worst_pct = dp
                worst_day = day

    # Course-wise breakdown
    course_stats = defaultdict(lambda: {"present": 0, "total": 0, "name": ""})
    for rec in records:
        sess = db.query(Session).filter(Session.id == rec.session_id).first()
        if sess:
            course = db.query(Course).filter(Course.id == sess.course_id).first()
            cname = course.name if course else f"Course {sess.course_id}"
            course_stats[sess.course_id]["name"] = cname
            course_stats[sess.course_id]["total"] += 1
            if rec.status.value == "present":
                course_stats[sess.course_id]["present"] += 1

    courses_breakdown = []
    for cid, stat in course_stats.items():
        cp = round((stat["present"] / stat["total"]) * 100) if stat["total"] else 0
        courses_breakdown.append({
            "course_id":   cid,
            "course_name": stat["name"],
            "present":     stat["present"],
            "total":       stat["total"],
            "percentage":  cp,
            "status":      "safe" if cp >= 75 else "warning" if cp >= 60 else "danger"
        })
    courses_breakdown.sort(key=lambda x: x["percentage"])

    # Generate insights
    insights = []
    if pct >= 75:
        insights.append({"type": "success", "message": f"Your attendance is {pct}% - above the 75% minimum. Keep it up!"})
    elif pct >= 60:
        insights.append({"type": "warning", "message": f"Your attendance is {pct}% - below 75%. You need {needed} more consecutive classes to be safe."})
    else:
        insights.append({"type": "danger", "message": f"Critical! Your attendance is only {pct}%. You need {needed} more classes to reach 75%."})

    if worst_day and worst_pct < 70:
        insights.append({"type": "warning", "message": f"You miss classes most on {worst_day}s ({worst_pct}% attendance). Try to improve."})

    danger_courses = [c for c in courses_breakdown if c["status"] == "danger"]
    if danger_courses:
        names = ", ".join(c["course_name"] for c in danger_courses[:2])
        insights.append({"type": "danger", "message": f"Critical shortage in: {names}. Attend every remaining class."})

    return {
        "student": {"id": student.id, "name": student.full_name, "inst_id": student.inst_id},
        "overall": {
            "present":    present_count,
            "total":      total_sessions,
            "percentage": pct,
            "status":     "safe" if pct >= 75 else "warning" if pct >= 60 else "danger",
            "classes_needed_for_75": needed if pct < 75 else 0,
        },
        "worst_day":         {"day": worst_day, "percentage": worst_pct} if worst_day else None,
        "courses_breakdown": courses_breakdown,
        "insights":          insights,
    }


@router.get("/insights/section")
def section_insights(
    branch:  Optional[str] = None,
    section: Optional[str] = None,
    current_user: User = Depends(require_roles(UserRole.faculty, UserRole.admin)),
    db: DBSession = Depends(get_db),
):
    """AI insights for an entire section — identify at-risk students."""
    q = db.query(User).filter(User.role == UserRole.student, User.status == "active")
    if branch:  q = q.filter(User.branch  == branch)
    if section: q = q.filter(User.section == section)
    students = q.all()

    total_sessions = db.query(Session).filter(
        Session.status == "closed",
        Session.branch  == branch,
        Session.section == section,
    ).count() if branch and section else 0

    at_risk, safe, critical = [], [], []
    for stu in students:
        records = db.query(AttendanceRecord)\
            .filter(AttendanceRecord.student_id == stu.id).all()
        present = sum(1 for r in records if r.status.value == "present")
        pct = round((present / total_sessions) * 100) if total_sessions else 0
        entry = {"id": stu.id, "name": stu.full_name, "inst_id": stu.inst_id,
                 "present": present, "total": total_sessions, "percentage": pct}
        if pct < 60:   critical.append(entry)
        elif pct < 75: at_risk.append(entry)
        else:          safe.append(entry)

    return {
        "section_summary": {
            "branch": branch, "section": section,
            "total_students":   len(students),
            "total_sessions":   total_sessions,
            "safe_count":       len(safe),
            "at_risk_count":    len(at_risk),
            "critical_count":   len(critical),
        },
        "critical_students": sorted(critical, key=lambda x: x["percentage"]),
        "at_risk_students":  sorted(at_risk,  key=lambda x: x["percentage"]),
        "safe_students":     sorted(safe,      key=lambda x: x["percentage"], reverse=True),
    }
