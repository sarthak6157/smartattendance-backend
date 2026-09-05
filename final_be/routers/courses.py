"""Courses routes — with bulk Excel upload and semester auto-detection."""
import io, re
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session as DBSession
from sqlalchemy import func

from core.security import get_current_user, require_roles
from db.database import get_db
from models.models import Course, User, UserRole
from schemas.schemas import CourseCreate, CourseOut

router   = APIRouter()
AdminOnly = require_roles(UserRole.admin)

DAYS = ["monday","tuesday","wednesday","thursday","friday","saturday"]

# ── Detect semester from course code ─────────────────────────────────────────
def detect_semester(code: str) -> Optional[str]:
    """
    Course code pattern: EAS211 → 4th char = '2' → 2nd semester
    ESA111 → 4th char = '1' → 1st semester
    """
    code = code.strip().upper()
    if len(code) >= 4:
        ch = code[3]
        if ch.isdigit() and ch != '0':
            return f"{ch}th" if int(ch) > 3 else ["","1st","2nd","3rd"][int(ch)]
    return None

def detect_year(code: str) -> Optional[str]:
    sem = detect_semester(code)
    if not sem: return None
    s = int(re.search(r'\d', sem).group())
    if s <= 2: return "1st Year"
    if s <= 4: return "2nd Year"
    if s <= 6: return "3rd Year"
    return "4th Year"

def detect_type(name: str) -> str:
    name = name.lower()
    if any(w in name for w in ["lab","practical","drawing","workshop"]): return "lab"
    return "theory"


# ── List all courses ──────────────────────────────────────────────────────────
@router.get("", response_model=list[CourseOut])
def list_courses(
    semester: Optional[str] = None,
    _: User = Depends(get_current_user),
    db: DBSession = Depends(get_db)
):
    q = db.query(Course).order_by(Course.semester, Course.code)
    if semester:
        q = q.filter(func.lower(Course.semester) == semester.lower())
    return q.all()


# ── Create single course ──────────────────────────────────────────────────────
@router.post("", response_model=CourseOut, status_code=201)
def create_course(
    payload: CourseCreate,
    _: User = Depends(AdminOnly),
    db: DBSession = Depends(get_db)
):
    code = payload.code.strip().upper()
    if db.query(Course).filter(Course.code == code).first():
        raise HTTPException(status_code=400, detail=f"Course code {code} already exists.")
    semester = payload.semester or detect_semester(code)
    course = Course(
        code=code, name=payload.name,
        department=payload.department,
        branch=payload.branch,
        section=payload.section,
        semester=semester,
        course_type=payload.course_type or detect_type(payload.name),
        credits=payload.credits if payload.credits is not None else 3,
    )
    db.add(course); db.commit(); db.refresh(course)
    return course


# ── Bulk upload via Excel/CSV ─────────────────────────────────────────────────
@router.post("/bulk-upload")
async def bulk_upload_courses(
    file: UploadFile = File(...),
    _: User = Depends(AdminOnly),
    db: DBSession = Depends(get_db),
):
    """
    Upload Excel/CSV with columns: Code, Name, Credits (optional)
    Semester is auto-detected from course code.
    """
    content = await file.read()
    ext = file.filename.lower().split('.')[-1]

    rows = []
    if ext == 'csv':
        import csv
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif ext in ('xlsx','xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else '') for i,v in enumerate(row)})
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed.")
    else:
        raise HTTPException(status_code=400, detail="Only .csv, .xlsx or .xls files supported.")

    added, skipped, errors = [], [], []
    for row in rows:
        # Flexible column name matching
        code = (row.get('Code') or row.get('code') or row.get('CODE') or '').strip().upper()
        name = (row.get('Name') or row.get('name') or row.get('Subject') or row.get('subject') or '').strip()
        credits = int(row.get('Credits') or row.get('credits') or 3)

        if not code or not name:
            continue

        if db.query(Course).filter(Course.code == code).first():
            skipped.append(code)
            continue

        semester = detect_semester(code)
        course   = Course(
            code=code, name=name,
            semester=semester,
            course_type=detect_type(name),
            credits=credits,
        )
        db.add(course)
        added.append(code)

    db.commit()
    return {
        "added":   len(added),
        "skipped": len(skipped),
        "message": f"Added {len(added)} courses. Skipped {len(skipped)} duplicates.",
        "added_codes":   added,
        "skipped_codes": skipped,
    }


# ── Download Excel template ───────────────────────────────────────────────────
@router.get("/template")
def download_template(_: User = Depends(AdminOnly)):
    """Download Excel template for bulk course upload."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses"

    # Header
    headers = ["Code", "Name", "Credits"]
    header_fill = PatternFill("solid", fgColor="1a3c6e")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample rows
    samples = [
        ("EAS211", "Engineering Mathematics-II", 4),
        ("EAS212", "Engineering Physics", 4),
        ("ECS201", "Computer Basics & C Programming", 3),
        ("ECS251", "C Programming Lab", 2),
        ("EEE217", "Basic Electrical Engineering", 3),
        ("TGE203", "English Communication-II", 2),
    ]
    for r, (code, name, cred) in enumerate(samples, 2):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=cred)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10

    # Note row
    note_row = len(samples) + 3
    ws.merge_cells(f"A{note_row}:C{note_row}")
    ws.cell(row=note_row, column=1,
            value="Note: Semester is auto-detected from course code (4th character). E.g. EAS211 → 2nd Semester").font = Font(italic=True, color="666666")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=courses_template.xlsx"}
    )


# ── Get courses grouped by semester ──────────────────────────────────────────
@router.get("/by-semester")
def courses_by_semester(_: User = Depends(get_current_user), db: DBSession = Depends(get_db)):
    """Return all courses grouped by semester — used by timetable builder."""
    courses = db.query(Course).order_by(Course.semester, Course.code).all()
    grouped = {}
    for c in courses:
        sem = c.semester or "Unknown"
        if sem not in grouped:
            grouped[sem] = []
        grouped[sem].append({
            "id": c.id, "code": c.code, "name": c.name,
            "credits": c.credits, "course_type": c.course_type,
        })
    return grouped


# ── Delete course ─────────────────────────────────────────────────────────────
@router.delete("/{course_id}", status_code=204)
def delete_course(course_id: str, _: User = Depends(AdminOnly), db: DBSession = Depends(get_db)):
    course = db.query(Course).filter(Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found.")
    db.delete(course); db.commit()


# ── Update course ─────────────────────────────────────────────────────────────
@router.patch("/{course_id}", response_model=CourseOut)
def update_course(
    course_id: str,
    payload: CourseCreate,
    _: User = Depends(AdminOnly),
    db: DBSession = Depends(get_db),
):
    course = db.query(Course).filter(Course.id == course_id).first()
    if not course: raise HTTPException(status_code=404, detail="Course not found.")
    course.name        = payload.name
    course.credits     = payload.credits if payload.credits is not None else course.credits
    course.course_type = payload.course_type or course.course_type
    db.commit(); db.refresh(course)
    return course
