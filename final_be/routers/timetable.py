"""Timetable routes — visual grid builder, copy, conflict detection."""
import re
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import func, or_
from sqlalchemy.orm import Session as DBSession

from core.security import get_current_user, require_roles
from db.database import get_db
from models.models import (TimetableSlot, Session, SessionStatus,
                           User, UserRole, Course, DayOfWeek)

router    = APIRouter()
AdminOnly = require_roles(UserRole.admin)

DAYS = ["monday","tuesday","wednesday","thursday","friday","saturday"]

# Default time slots
DEFAULT_SLOTS = [
    "09:10-09:25",
    "09:30-10:30",
    "10:30-11:25",
    "11:30-12:25",
    "12:30-13:25",
    "14:25-15:10",
    "15:10-15:55",
]

# ── Pydantic schemas ──────────────────────────────────────────────────────────
class SlotCreate(BaseModel):
    course_id:   int
    faculty_id:  Optional[str] = None   # optional for free classes
    day_of_week: str
    start_time:  str
    end_time:    str
    room:        Optional[str] = None
    branch:      Optional[str] = None
    section:     Optional[str] = None
    sub_section: Optional[str] = None
    semester:    Optional[str] = None
    course_type: Optional[str] = None

class SlotOut(BaseModel):
    id:           int
    course_id:    int
    faculty_id:   Optional[str] = None
    day_of_week:  str
    start_time:   str
    end_time:     str
    room:         Optional[str] = None
    branch:       Optional[str] = None
    section:      Optional[str] = None
    sub_section:  Optional[str] = None
    semester:     Optional[str] = None
    course_type:  Optional[str] = None
    is_active:    bool
    course_name:  Optional[str] = None
    course_code:  Optional[str] = None
    faculty_name: Optional[str] = None
    model_config  = {"from_attributes": True, "use_enum_values": True}

class GoLiveRequest(BaseModel):
    gps_lat: Optional[str] = None
    gps_lng: Optional[str] = None

class CopyTimetableRequest(BaseModel):
    from_branch:   str
    from_section:  str
    from_semester: str
    to_branch:     str
    to_section:    str
    to_semester:   str
    copy_teachers: bool = False  # if False, teachers = unassigned (faculty_id=1)


# ── Helper: extract core branch for fuzzy match ───────────────────────────────
def extract_core(branch_str: str) -> str:
    if not branch_str: return ""
    s = branch_str.strip().lower()
    for p in ["b.tech - ","b.tech-","b.tech ","btech ","b.e - ","b.e ","m.tech - ","m.tech "]:
        if s.startswith(p): s = s[len(p):]; break
    return re.sub(r"[\s\-_]+", " ", s).strip()


# ── List slots ────────────────────────────────────────────────────────────────
@router.get("", response_model=list[SlotOut])
def list_slots(
    branch:     Optional[str] = Query(None),
    section:    Optional[str] = Query(None),
    semester:   Optional[str] = Query(None),
    faculty_id: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    q = db.query(TimetableSlot).filter(TimetableSlot.is_active == True)

    if current_user.role == UserRole.faculty:
        # Include unassigned "free period" slots (Library, Coding Practice,
        # etc.) too — these have no owning teacher, but any faculty needs to
        # be able to see and go-live them, otherwise attendance can never be
        # taken for a free period at all.
        q = q.filter(or_(TimetableSlot.faculty_id == current_user.id, TimetableSlot.faculty_id == None))
    elif faculty_id:
        q = q.filter(TimetableSlot.faculty_id == faculty_id)

    if current_user.role == UserRole.student:
        effective_branch   = branch  or current_user.branch or current_user.department
        effective_section  = section or current_user.section
        # Lab batch (e.g. "C1") lives in User.sub_section, NOT User.course
        # (course holds the degree type, e.g. "B.Tech"). Reading the wrong
        # field here meant effective_subsection was almost always empty,
        # which fell through to showing every batch's lab slots.
        effective_subsection = (current_user.sub_section or "").strip() or None
    else:
        effective_branch     = branch
        effective_section    = section
        effective_subsection = None

    if effective_branch:
        eb      = effective_branch.strip().lower()
        eb_core = extract_core(eb)
        eb_short = eb_core.split("(")[0].strip() if eb_core else ""
        conditions = [
            TimetableSlot.branch == None,
            TimetableSlot.branch == "",
            func.lower(TimetableSlot.branch) == eb,
            func.strpos(func.lower(TimetableSlot.branch), eb) > 0,
            func.strpos(eb, func.lower(TimetableSlot.branch)) > 0,
        ]
        if eb_core:  conditions.append(func.strpos(func.lower(TimetableSlot.branch), eb_core) > 0)
        if eb_short: conditions.append(func.strpos(func.lower(TimetableSlot.branch), eb_short) > 0)
        q = q.filter(or_(*conditions))

    if effective_section:
        sec = effective_section.strip().upper()
        if current_user.role == UserRole.student and effective_subsection:
            subsec = effective_subsection.strip().upper()
            q = q.filter(
                func.lower(TimetableSlot.section) == sec.lower(),
                or_(
                    TimetableSlot.sub_section == None,
                    func.lower(TimetableSlot.sub_section) == subsec.lower()
                )
            )
        else:
            q = q.filter(func.lower(TimetableSlot.section) == sec.lower())

    if semester:
        q = q.filter(func.lower(TimetableSlot.semester) == semester.strip().lower())

    slots = q.order_by(TimetableSlot.day_of_week, TimetableSlot.start_time).all()
    if not slots: return []

    # Bulk load courses and faculty
    course_ids  = list({s.course_id  for s in slots})
    faculty_ids = list({s.faculty_id for s in slots})
    courses_map = {c.id: c for c in db.query(Course).filter(Course.id.in_(course_ids)).all()}
    faculty_map = {f.id: f for f in db.query(User).filter(User.id.in_(faculty_ids)).all()}

    result = []
    for s in slots:
        co  = courses_map.get(s.course_id)
        fac = faculty_map.get(s.faculty_id)
        d   = SlotOut.model_validate(s)
        d.day_of_week  = s.day_of_week.value if hasattr(s.day_of_week, "value") else str(s.day_of_week)
        d.course_name  = co.name       if co  else None
        d.course_code  = co.code       if co  else None
        d.faculty_name = fac.full_name if fac else None
        result.append(d)
    return result


# ── Get grid data for visual timetable builder ────────────────────────────────
@router.get("/grid")
def get_timetable_grid(
    branch:   str = Query(...),
    section:  str = Query(...),
    semester: str = Query(...),
    _: User = Depends(require_roles(UserRole.admin)),
    db: DBSession = Depends(get_db),
):
    """
    Returns timetable as a grid dict:
    { "monday": { "09:30-10:30": { slot_data }, ... }, ... }
    Used by the visual timetable builder.
    """
    slots = db.query(TimetableSlot).filter(
        TimetableSlot.is_active == True,
        func.lower(TimetableSlot.branch)   == branch.strip().lower(),
        func.lower(TimetableSlot.section)  == section.strip().lower(),
        func.lower(TimetableSlot.semester) == semester.strip().lower(),
    ).all()

    course_ids  = list({s.course_id  for s in slots})
    faculty_ids = list({s.faculty_id for s in slots})
    courses_map = {c.id: c for c in db.query(Course).filter(Course.id.in_(course_ids)).all()} if course_ids else {}
    faculty_map = {f.id: f for f in db.query(User).filter(User.id.in_(faculty_ids)).all()} if faculty_ids else {}

    grid = {day: {} for day in DAYS}
    for s in slots:
        day = s.day_of_week.value if hasattr(s.day_of_week, "value") else str(s.day_of_week)
        time_key = f"{s.start_time}-{s.end_time}"
        co  = courses_map.get(s.course_id)
        fac = faculty_map.get(s.faculty_id)
        entry = {
            "id":           s.id,
            "course_id":    s.course_id,
            "course_name":  co.name  if co  else "?",
            "course_code":  co.code  if co  else "?",
            "course_type":  s.course_type or (co.course_type if co else "theory"),
            "faculty_id":   s.faculty_id,
            "faculty_name": fac.full_name if fac else "Unassigned",
            "room":         s.room,
            "sub_section":  s.sub_section,
        }
        # A cell can legitimately hold more than one slot — e.g. two lab
        # batches (C1, C2) running in parallel at the same day/time. Store a
        # LIST per cell so none of them get silently overwritten.
        grid[day].setdefault(time_key, []).append(entry)
    return {"grid": grid, "time_slots": DEFAULT_SLOTS, "days": DAYS}


# ── Create slot ───────────────────────────────────────────────────────────────
@router.post("", response_model=SlotOut, status_code=201)
def create_slot(
    payload: SlotCreate,
    _: User = Depends(AdminOnly),
    db: DBSession = Depends(get_db),
):
    co  = db.query(Course).filter(Course.id == payload.course_id).first()
    if not co:  raise HTTPException(status_code=404, detail="Course not found.")

    # Free classes (Library, Tinkerer, Mentor) don't need a teacher
    FREE_CODES = {"LIBRARY", "TINKERER", "MENTOR", "CODING", "FREE"}
    is_free = co.code.upper() in FREE_CODES or co.credits == 0

    fac = None
    if payload.faculty_id:
        fac = db.query(User).filter(User.id == payload.faculty_id, User.role == UserRole.faculty).first()
        if not fac:
            raise HTTPException(status_code=404, detail="Faculty not found.")
    elif not is_free:
        raise HTTPException(status_code=400, detail="Please assign a teacher for this class.")

    data = payload.model_dump()
    data["day_of_week"] = data["day_of_week"].strip().lower()

    slot = TimetableSlot(**data)
    db.add(slot); db.commit(); db.refresh(slot)
    d = SlotOut.model_validate(slot)
    d.day_of_week  = slot.day_of_week.value if hasattr(slot.day_of_week, "value") else str(slot.day_of_week)
    d.course_name  = co.name
    d.course_code  = co.code
    d.faculty_name = fac.full_name if fac else None
    return d


# ── Update slot ───────────────────────────────────────────────────────────────
@router.patch("/{slot_id}", response_model=SlotOut)
def update_slot(
    slot_id: int,
    payload: SlotCreate,
    _: User = Depends(AdminOnly),
    db: DBSession = Depends(get_db),
):
    slot = db.query(TimetableSlot).filter(TimetableSlot.id == slot_id).first()
    if not slot: raise HTTPException(status_code=404, detail="Slot not found.")
    co  = db.query(Course).filter(Course.id == payload.course_id).first()
    if not co:  raise HTTPException(status_code=404, detail="Course not found.")

    # Free classes (Library, Tinkerer, Mentor) don't need a teacher — same rule as create_slot
    FREE_CODES = {"LIBRARY", "TINKERER", "MENTOR", "CODING", "FREE"}
    is_free = co.code.upper() in FREE_CODES or co.credits == 0

    fac = None
    if payload.faculty_id:
        fac = db.query(User).filter(User.id == payload.faculty_id, User.role == UserRole.faculty).first()
        if not fac:
            raise HTTPException(status_code=404, detail="Faculty not found.")
    elif not is_free:
        raise HTTPException(status_code=400, detail="Please assign a teacher for this class.")

    for k, v in payload.model_dump().items():
        if k == "day_of_week": v = v.strip().lower()
        setattr(slot, k, v)
    db.commit(); db.refresh(slot)

    d = SlotOut.model_validate(slot)
    d.day_of_week  = slot.day_of_week.value if hasattr(slot.day_of_week, "value") else str(slot.day_of_week)
    d.course_name  = co.name
    d.course_code  = co.code
    d.faculty_name = fac.full_name if fac else None
    return d


# ── Delete slot ───────────────────────────────────────────────────────────────
@router.delete("/{slot_id}", status_code=204)
def delete_slot(slot_id: int, _: User = Depends(AdminOnly), db: DBSession = Depends(get_db)):
    slot = db.query(TimetableSlot).filter(TimetableSlot.id == slot_id).first()
    if not slot: raise HTTPException(status_code=404, detail="Slot not found.")
    db.delete(slot); db.commit()


# ── Copy timetable ────────────────────────────────────────────────────────────
@router.post("/copy")
def copy_timetable(
    payload: CopyTimetableRequest,
    _: User = Depends(AdminOnly),
    db: DBSession = Depends(get_db),
):
    """Copy all slots from one section to another. Optionally copy teachers."""
    slots = db.query(TimetableSlot).filter(
        TimetableSlot.is_active == True,
        func.lower(TimetableSlot.branch)   == payload.from_branch.strip().lower(),
        func.lower(TimetableSlot.section)  == payload.from_section.strip().lower(),
        func.lower(TimetableSlot.semester) == payload.from_semester.strip().lower(),
    ).all()

    if not slots:
        raise HTTPException(status_code=404, detail="No slots found for source section.")

    # Delete existing slots in destination if any
    db.query(TimetableSlot).filter(
        TimetableSlot.is_active == True,
        func.lower(TimetableSlot.branch)   == payload.to_branch.strip().lower(),
        func.lower(TimetableSlot.section)  == payload.to_section.strip().lower(),
        func.lower(TimetableSlot.semester) == payload.to_semester.strip().lower(),
    ).delete(synchronize_session=False)

    # Admin user id=1 as placeholder for unassigned
    admin = db.query(User).filter(User.role == UserRole.admin).first()
    placeholder_id = admin.id if admin else None

    new_slots = []
    for s in slots:
        new_slot = TimetableSlot(
            course_id   = s.course_id,
            faculty_id  = s.faculty_id if payload.copy_teachers else placeholder_id,
            day_of_week = s.day_of_week,
            start_time  = s.start_time,
            end_time    = s.end_time,
            room        = s.room,
            branch      = payload.to_branch,
            section     = payload.to_section,
            sub_section = s.sub_section,
            semester    = payload.to_semester,
            course_type = s.course_type,
            is_active   = True,
        )
        db.add(new_slot)
        new_slots.append(new_slot)

    db.commit()
    return {
        "copied":  len(new_slots),
        "message": f"Copied {len(new_slots)} slots to {payload.to_branch} Section {payload.to_section}.",
        "note":    "Teachers are unassigned — please assign them in the timetable editor." if not payload.copy_teachers else "Teachers copied from source section.",
    }


# ── Conflict check ────────────────────────────────────────────────────────────
@router.get("/conflicts")
def check_conflicts(
    branch:   Optional[str] = None,
    section:  Optional[str] = None,
    semester: Optional[str] = None,
    _: User = Depends(AdminOnly),
    db: DBSession = Depends(get_db),
):
    """Check for teacher conflicts — same teacher at same time in different places."""
    q = db.query(TimetableSlot).filter(TimetableSlot.is_active == True)
    if branch:   q = q.filter(func.lower(TimetableSlot.branch)   == branch.strip().lower())
    if section:  q = q.filter(func.lower(TimetableSlot.section)  == section.strip().lower())
    if semester: q = q.filter(func.lower(TimetableSlot.semester) == semester.strip().lower())
    slots = q.all()

    faculty_ids = list({s.faculty_id for s in slots})
    faculty_map = {f.id: f for f in db.query(User).filter(User.id.in_(faculty_ids)).all()}
    course_ids  = list({s.course_id  for s in slots})
    courses_map = {c.id: c for c in db.query(Course).filter(Course.id.in_(course_ids)).all()}

    # Group by faculty + day + time
    from collections import defaultdict
    schedule = defaultdict(list)
    for s in slots:
        day = s.day_of_week.value if hasattr(s.day_of_week, "value") else str(s.day_of_week)
        key = (s.faculty_id, day, s.start_time)
        schedule[key].append(s)

    conflicts = []
    for (fac_id, day, time), slot_list in schedule.items():
        if len(slot_list) > 1:
            fac = faculty_map.get(fac_id)
            conflicts.append({
                "teacher":    fac.full_name if fac else f"Faculty {fac_id}",
                "day":        day,
                "time":       time,
                "sections":   [f"Sec {s.section}" for s in slot_list],
                "subjects":   [courses_map.get(s.course_id, type('x', (), {'name':'?'})()).name for s in slot_list],
            })

    return {"conflicts": conflicts, "total": len(conflicts)}


# ── Go Live ───────────────────────────────────────────────────────────────────
@router.post("/{slot_id}/go-live")
def go_live(
    slot_id: int,
    payload: GoLiveRequest,
    current_user: User = Depends(require_roles(UserRole.faculty)),
    db: DBSession = Depends(get_db),
):
    import secrets
    slot = db.query(TimetableSlot).filter(TimetableSlot.id == slot_id).first()
    if not slot: raise HTTPException(status_code=404, detail="Slot not found.")
    # Free-period slots (Library, Coding Practice, etc.) have no assigned
    # teacher — any faculty can go live for them. Assigned slots still
    # require the owning faculty.
    if slot.faculty_id is not None and slot.faculty_id != current_user.id:
        raise HTTPException(status_code=403, detail="This slot is not assigned to you.")

    active = db.query(Session).filter(
        Session.faculty_id == current_user.id,
        Session.status     == SessionStatus.active,
    ).first()
    if active:
        raise HTTPException(status_code=400, detail="You already have an active session. End it first.")

    co  = db.query(Course).filter(Course.id == slot.course_id).first()
    now = datetime.utcnow()

    try:
        import pytz
        ist = pytz.timezone("Asia/Kolkata")
        ist_now = datetime.now(ist).replace(tzinfo=None)
    except ImportError:
        from datetime import timedelta
        ist_now = now + timedelta(hours=5, minutes=30)

    day_name = DAYS[ist_now.weekday()] if ist_now.weekday() < 6 else "saturday"
    slot_day = slot.day_of_week.value if hasattr(slot.day_of_week, "value") else str(slot.day_of_week)

    if slot_day.lower() != day_name.lower():
        raise HTTPException(status_code=400, detail=f"This slot is for {slot_day.title()}, not today ({day_name.title()}).")

    sh, sm = map(int, slot.start_time.split(":"))
    eh, em = map(int, slot.end_time.split(":"))
    now_m  = ist_now.hour * 60 + ist_now.minute
    slot_start_m = sh * 60 + sm
    slot_end_m   = eh * 60 + em

    if now_m < slot_start_m - 10:
        raise HTTPException(status_code=400, detail=f"Too early! Class starts at {slot.start_time}.")
    if now_m > slot_end_m:
        raise HTTPException(status_code=400, detail="Class time has passed.")

    session = Session(
        course_id    = slot.course_id,
        # Session.faculty_id is required (attendance/audit trail needs a
        # responsible person on record) — for a free period, that's whoever
        # actually went live, not the slot's (missing) assigned teacher.
        faculty_id   = slot.faculty_id if slot.faculty_id is not None else current_user.id,
        timetable_id = slot.id,
        title        = f"{co.name if co else 'Class'} - {slot.section} {slot.start_time}",
        location     = slot.room,
        branch       = slot.branch,
        section      = slot.section,
        sub_section  = slot.sub_section,
        semester     = slot.semester,
        course_type  = slot.course_type,
        gps_lat      = payload.gps_lat,
        gps_lng      = payload.gps_lng,
        status       = SessionStatus.active,
        scheduled_at = now,
        started_at   = now,
        qr_token     = secrets.token_urlsafe(16),
        grace_minutes= 15,
    )
    db.add(session); db.commit(); db.refresh(session)
    return {
        "session_id": session.id,
        "qr_token":   session.qr_token,
        "title":      session.title,
        "message":    "Session is now LIVE!",
    }


# ── Bulk Timetable Import ─────────────────────────────────────────────────────
@router.get("/bulk-template")
def download_timetable_template(_: User = Depends(AdminOnly)):
    """Download Excel template for bulk timetable import."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    import io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"

    header_fill = PatternFill("solid", fgColor="1a3c6e")
    headers = ["Branch","Section","Semester","Day","StartTime","EndTime","CourseCode","FacultyID","Room","SubSection","CourseType"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample rows
    samples = [
        ("CSE (AI-ML-DL)","A","2nd","monday","09:30","10:30","EAS211","101","Room 301","","theory"),
        ("CSE (AI-ML-DL)","A","2nd","monday","10:30","11:25","TGE203","102","Room 302","","theory"),
        ("CSE (AI-ML-DL)","A","2nd","tuesday","11:30","12:25","ECS251","103","Lab 201","A1","lab"),
        ("CSE (AI-ML-DL)","A","2nd","tuesday","11:30","12:25","ECS251","104","Lab 202","A2","lab"),
        ("CSE (AI-ML-DL)","A","2nd","wednesday","09:10","09:25","LIBRARY","","Library","","theory"),
    ]
    for r, row in enumerate(samples, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)

    col_widths = [20,8,8,12,10,10,14,10,12,12,10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Notes
    note_row = len(samples) + 3
    ws.merge_cells(f"A{note_row}:K{note_row}")
    ws.cell(row=note_row, column=1, value="Notes: Day must be lowercase (monday/tuesday etc). FacultyID is the numeric ID from users table. Leave FacultyID blank for free classes (Library/Tinkerer/Mentor). SubSection is optional for labs (A1/A2).").font = Font(italic=True, color="666666")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=timetable_template.xlsx"}
    )


@router.post("/bulk-import")
async def bulk_import_timetable(
    file: UploadFile, File = File(...),
    clear_existing: bool = False,
    _: User = Depends(AdminOnly),
    db: DBSession = Depends(get_db),
):
    """
    Bulk import timetable from Excel/CSV.
    Columns: Branch,Section,Semester,Day,StartTime,EndTime,CourseCode,FacultyID,Room,SubSection,CourseType
    """
    import io as _io
    content = await file.read()
    ext = file.filename.lower().split('.')[-1]

    rows = []
    if ext == 'csv':
        import csv
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(_io.StringIO(text))
        rows = list(reader)
    elif ext in ('xlsx','xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(content))
            ws = wb.active
            headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v for v in row):  # skip empty rows
                    rows.append({headers[i]: (str(v).strip() if v is not None else '') for i,v in enumerate(row)})
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed.")
    else:
        raise HTTPException(status_code=400, detail="Only .csv, .xlsx or .xls files supported.")

    # Load all courses and faculty for lookup
    all_courses = {c.code.upper(): c for c in db.query(Course).all()}
    all_faculty  = {str(f.id): f for f in db.query(User).filter(User.role == UserRole.faculty).all()}
    FREE_CODES   = {"LIBRARY","TINKERER","MENTOR","CODING","FREE"}

    added, skipped, errors = [], [], []

    if clear_existing:
        # Get unique branch/section/semester combos from file and clear those
        combos = set()
        for row in rows:
            b = (row.get('Branch') or '').strip()
            s = (row.get('Section') or '').strip()
            sem = (row.get('Semester') or '').strip()
            if b and s and sem:
                combos.add((b.lower(), s.lower(), sem.lower()))
        for (b,s,sem) in combos:
            db.query(TimetableSlot).filter(
                func.lower(TimetableSlot.branch)   == b,
                func.lower(TimetableSlot.section)  == s,
                func.lower(TimetableSlot.semester) == sem,
                TimetableSlot.is_active == True,
            ).delete(synchronize_session=False)

    for i, row in enumerate(rows, 2):
        try:
            branch     = (row.get('Branch')     or '').strip()
            section    = (row.get('Section')    or '').strip()
            semester   = (row.get('Semester')   or '').strip()
            day        = (row.get('Day')        or '').strip().lower()
            start_time = (row.get('StartTime')  or '').strip()
            end_time   = (row.get('EndTime')    or '').strip()
            course_code= (row.get('CourseCode') or '').strip().upper()
            faculty_id = (row.get('FacultyID')  or '').strip()
            room       = (row.get('Room')       or '').strip() or None
            sub_section= (row.get('SubSection') or '').strip().upper() or None
            course_type= (row.get('CourseType') or 'theory').strip().lower()

            if not all([branch, section, semester, day, start_time, end_time, course_code]):
                errors.append(f"Row {i}: Missing required fields")
                continue

            if day not in DAYS:
                errors.append(f"Row {i}: Invalid day '{day}' — must be monday/tuesday etc.")
                continue

            course = all_courses.get(course_code)
            if not course:
                errors.append(f"Row {i}: Course code '{course_code}' not found — add it in Courses first.")
                continue

            is_free = course_code in FREE_CODES or (course.credits is not None and course.credits == 0)

            fac_id = None
            if faculty_id:
                if faculty_id not in all_faculty:
                    errors.append(f"Row {i}: Faculty ID '{faculty_id}' not found.")
                    continue
                fac_id = int(faculty_id)
            elif not is_free:
                errors.append(f"Row {i}: Faculty ID required for non-free class '{course_code}'.")
                continue

            slot = TimetableSlot(
                course_id   = course.id,
                faculty_id  = fac_id,
                day_of_week = day,
                start_time  = start_time,
                end_time    = end_time,
                room        = room,
                branch      = branch,
                section     = section,
                sub_section = sub_section,
                semester    = semester,
                course_type = course_type,
                is_active   = True,
            )
            db.add(slot)
            added.append(f"{day} {start_time} {course_code}")

        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()

    return {
        "added":   len(added),
        "skipped": len(skipped),
        "errors":  errors[:20],  # show max 20 errors
        "message": f"Added {len(added)} slots. {len(errors)} errors.",
    }
